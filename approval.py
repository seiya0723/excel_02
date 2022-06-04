import openpyxl as px
import time 

old_data    = None

#担当、課長、部長の承認状態をチェックする( 前のループと比較して不一致(何か記入された場合)の時メールを送信する )
while True:

    wb          = px.load_workbook('approval.xlsx')
    ws          = wb.active
    
    rows        = ws.iter_rows()
    row_number  = 0

    new_data    = []
    
    #行ごとにループさせる。
    for row in rows:
    
        row_number += 1
    
        if row_number == 1:
            #最初のループはヘッダなので、スキップする。
            continue
    
        cell_number = 0
    
        dic = {}
    
        dic["企画名"]   = row[0].value
        dic["担当"]     = row[1].value
        dic["課長"]     = row[2].value
        dic["部長"]     = row[3].value
        dic["B社"]      = row[4].value
    
        #現在の状況を記録
        new_data.append(dic)
        print(new_data)
    
    
    #最初のループの時(old_dataがNoneの時)、現在のデータを入れ、次のループへ
    if not old_data:

        old_data = new_data
        time.sleep(3)

        continue
    
    
    #現在のデータと過去データを比較する(相違点があれば、メールする)
    #TIPS:複数のリストを同時にループする時、zipを使う。
    for new_row,old_row in zip(new_data,old_data):
    
        print(new_row["企画名"])
    
        if new_row["担当"] != old_row["担当"]:
            print("担当の承認状況が変わりました。")
    
        if new_row["課長"] != old_row["課長"]:
            print("課長の承認状況が変わりました。")
    
        if new_row["部長"] != old_row["部長"]:
            print("部長の承認状況が変わりました。")
    
        if new_row["B社"] != old_row["B社"]:
            print("B社の承認状況が変わりました。")
    
    old_data = new_data
    time.sleep(3)



import sendgrid
from sendgrid.helpers.mail import *

SENDGRID_API    = "ここにSendgridのAPIキーを"
SG              = sendgrid.SendGridAPIClient(api_key=SENDGRID_API)

def send_mail(to_email,from_email,subject,message):
    
    mail        = Mail( Email(from_email),
                    To(to_email),
                    subject,
                    Content("text/plain", message),
                    )
            


#未確認、未承認と承認をチェックする(セルをチェックする時、未確認、未承認と承認を文字列で判定する)
"""

#whileループで前のシートの状態を記録し、比較する

rows    = ws.iter_rows()

#行ごとにループさせる。
for row in rows:
    for cell in row:

        if type(cell.value) == type(str):

            if cell.value == "承認"
                #承認
            else:
                #未承認

        else:
            #値が入っているかどうかのチェック(Noneであれば未確認)
            if cell.value == None:
                #未確認
            else:
                #未承認


#この時に新たに承認が確認されれば、メールを送信する。
"""
